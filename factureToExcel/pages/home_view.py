from typing import Union
import flet as ft
from pages.Router import Router, DataStrategyEnum

import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import os
import time
def get_unique_sheet_name(workbook, base_name):
    sheet_name = base_name
    count = 1
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{base_name}_{count}"
        count += 1
    return sheet_name
def HomeView(page: Router):

    def generate_csv_from_xml(e: ft.FilePickerResultEvent):
        try:
            if e.files is None:
                return
            excel_path = f'facturas_{time.strftime("%m%d")}.xlsx'

            if os.path.exists(excel_path):
                count = 1
                while os.path.exists(f'facturas_{time.strftime("%m%d")}_{count}.xlsx'):
                    count += 1
                excel_path = f'facturas_{time.strftime("%m%d")}_{count}.xlsx'
                pd.DataFrame().to_excel(excel_path, index=False, engine='openpyxl')
            else:
                pd.DataFrame().to_excel(excel_path, index=False, engine='openpyxl')

            for file in e.files:
                file = file.path
                tree = ET.parse(file)
                root = tree.getroot()

                namespaces = {
                    'cfdi': 'http://www.sat.gob.mx/cfd/4'
                }

                conceptos = root.findall(".//cfdi:Concepto", namespaces)
                data = []
                emisor = root.find(".//cfdi:Emisor", namespaces)
                nombre_emisor = emisor.attrib.get("Nombre")

                for concepto in conceptos:
                    descripcion = concepto.attrib.get("Descripcion")
                    precio_unitario = concepto.attrib.get("ValorUnitario")
                    descuento = concepto.attrib.get("Descuento")
                    if descuento is not None:
                        importe = concepto.attrib.get("Importe")
                        total_with_discount = float(importe) - float(descuento)
                        precio_unitario = total_with_discount / float(concepto.attrib.get("Cantidad"))
                    if descripcion is not None:
                        descripcion = descripcion.upper()
                        descripcion = descripcion.replace("Á", "A").replace("É", "E").replace("Í", "I")
                        descripcion = descripcion.replace("Ó", "O").replace("Ú", "U")
                    data.append({
                        "ClaveProdServ": concepto.attrib.get("ClaveProdServ"),
                        "Cantidad": concepto.attrib.get("Cantidad"),
                        "Unidad": concepto.attrib.get("Unidad") if concepto.attrib.get("Unidad") else "PIEZA",
                        "Descripcion": descripcion,
                        "Precio unitario": precio_unitario,
                    })

                df = pd.DataFrame(data)

                workbook = load_workbook(excel_path)
                
                sheet_name = get_unique_sheet_name(workbook, nombre_emisor)

                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

                workbook = load_workbook(excel_path)
                
                worksheet = workbook[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                for row in range(2, worksheet.max_row + 1):
                    worksheet[f'E{row}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

                workbook.save(excel_path)
                alert = ft.AlertDialog(
                    modal=True,
                    title=ft.Text("Archivo generado correctamente"),
                    content=ft.Text(f'Archivo generado en {excel_path}'),
                    actions=[
                        ft.TextButton("Cerrar", on_click=lambda e: page.close_dialog())
                    ],
                    actions_alignment=ft.MainAxisAlignment.END,
                    open=True
                )


                page.set_dialog(alert)
                page.update()     
        
            workbook = load_workbook(excel_path)
            workbook.remove(workbook['Sheet1'])
            workbook.save(excel_path)
            workbook.close()
            os.startfile(excel_path)
        except Exception as e:
            alert = ft.AlertDialog(
                            modal=True,
                            title=ft.Text("Error"),
                            content=ft.Text(f'Error al generar el archivo: {e}'),
                            actions=[
                                ft.TextButton("Cerrar", on_click=lambda e: page.close_dialog())
                            ],
                            actions_alignment=ft.MainAxisAlignment.END,
                            open=True
                        )
            page.set_dialog(alert)
    pick_files_dialog = ft.FilePicker(on_result=generate_csv_from_xml)
    page.set_overlay(pick_files_dialog)
    content = ft.Column(
        [
            ft.Row([
                ft.Column([
                    ft.ElevatedButton(
                                        "Seleccionar archivo",
                                        icon=ft.icons.UPLOAD_FILE,
                                        on_click=lambda _: pick_files_dialog.pick_files(
                                            allow_multiple=True,
                                            allowed_extensions=[
                                                "xml"]
                                        ),
                                    adaptive=True,
                                    ),
                ft.Text("Selecciona un archivo XML para convertirlo a CSV", 
                        text_align=ft.TextAlign.CENTER),
                ], 
                alignment=ft.MainAxisAlignment.CENTER, 
                horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            ], alignment=ft.MainAxisAlignment.CENTER),

        ], 
        alignment=ft.MainAxisAlignment.CENTER, 
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        expand=True
    )

    return content


